"""
超简单的跟踪验证工具

逻辑：
1. 从 Excel 读取 Time 和 Signature
2. 用 Signature 查询链上时间
3. 筛选延迟 < 60秒的交易
4. 统计延迟分布
"""

import openpyxl
from datetime import datetime
import time
import requests
import os
from dotenv import load_dotenv

# 加载 .env 文件
load_dotenv()


class SimpleTrackingValidator:
    """简化的跟踪验证器"""
    
    def __init__(self, excel_file, api_key):
        """
        初始化
        
        参数:
            excel_file: Excel文件路径
            api_key: Helius API Key
        """
        self.excel_file = excel_file
        self.api_key = api_key
        self.REALTIME_THRESHOLD = 60  # 60秒阈值（超过就不是实时）
    
    def get_tx_timestamp(self, signature):
        """
        从 Helius API 获取交易链上时间
        
        参数:
            signature: 交易签名
        
        返回:
            timestamp: Unix时间戳（秒）
        """
        # 使用正确的 Helius API 端点（POST 方法）
        url = f"https://api.helius.xyz/v0/transactions/?api-key={self.api_key}"
        
        try:
            response = requests.post(url, json={"transactions": [signature]}, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                # Helius 返回的是数组，取第一个元素
                if data and isinstance(data, list) and len(data) > 0:
                    # blockTime 是 Unix 时间戳
                    return data[0].get('timestamp')
                else:
                    print(f"⚠️ 响应格式错误: {signature[:16]}...")
                    return None
            else:
                print(f"⚠️ API错误: {signature[:16]}... - {response.status_code}")
                return None
                
        except Exception as e:
            print(f"⚠️ 网络错误: {signature[:16]}... - {e}")
            return None
    
    def load_transactions(self):
        """
        从 Excel 加载交易
        
        返回:
            list: 交易列表
                [
                    {
                        'time_str': '2026-01-27 19:29:10',
                        'captured_at': 1738003750,
                        'signature': 'xxx...',
                        'timestamp': 1738003745,  # 从API查询
                        'delay': 5
                    }
                ]
        """
        print(f"📂 加载 Excel: {self.excel_file}")
        
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active
        
        transactions = []
        total_rows = ws.max_row - 1  # 减去表头
        
        print(f"📊 总交易数: {total_rows}笔")
        print(f"🔍 开始查询链上时间...\n")
        
        # 从第2行开始（跳过表头）
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            time_str = row[0]  # Time 列
            signature = row[1]  # Signature 列
            
            # 转换捕获时间为时间戳
            try:
                captured_at = datetime.strptime(str(time_str), "%Y-%m-%d %H:%M:%S").timestamp()
            except:
                print(f"⚠️ 跳过: 时间格式错误 - {time_str}")
                continue
            
            # 查询链上时间
            print(f"[{idx}/{total_rows}] 查询: {signature[:16]}...", end=" ")
            timestamp = self.get_tx_timestamp(signature)
            
            if timestamp is None:
                print("❌ 失败")
                continue
            
            # 计算延迟
            delay = captured_at - timestamp
            
            print(f"✅ 延迟: {delay:.1f}秒")
            
            transactions.append({
                'time_str': time_str,
                'captured_at': captured_at,
                'signature': signature,
                'timestamp': timestamp,
                'delay': delay
            })
            
            # 简单限流（避免API限速）
            time.sleep(0.1)
        
        print(f"\n✅ 加载完成: {len(transactions)}/{total_rows}笔\n")
        
        return transactions
    
    def filter_realtime(self, transactions):
        """
        筛选实时交易（延迟 < 60秒）
        
        返回:
            (realtime_txs, backfill_txs)
        """
        realtime_txs = []
        backfill_txs = []
        
        for tx in transactions:
            if tx['delay'] <= self.REALTIME_THRESHOLD:
                realtime_txs.append(tx)
            else:
                backfill_txs.append(tx)
        
        return realtime_txs, backfill_txs
    
    def calculate_stats(self, transactions):
        """
        计算延迟统计
        
        返回:
            dict: 统计结果
        """
        if not transactions:
            return None
        
        delays = [tx['delay'] for tx in transactions]
        delays.sort()
        
        n = len(delays)
        
        # 延迟分布
        distribution = {
            '0-5s': sum(1 for d in delays if 0 <= d < 5),
            '5-10s': sum(1 for d in delays if 5 <= d < 10),
            '10-15s': sum(1 for d in delays if 10 <= d < 15),
            '15-30s': sum(1 for d in delays if 15 <= d < 30),
            '30-60s': sum(1 for d in delays if 30 <= d <= 60)
        }
        
        return {
            'count': n,
            'avg_delay': sum(delays) / n,
            'median_delay': delays[n // 2],
            'min_delay': min(delays),
            'max_delay': max(delays),
            'p95_delay': delays[int(n * 0.95)] if n > 0 else 0,
            'distribution': distribution
        }
    
    def validate(self):
        """
        执行验证
        
        返回:
            dict: 验证报告
        """
        # 1. 加载交易
        all_txs = self.load_transactions()
        
        if not all_txs:
            return {'error': '没有可用的交易数据'}
        
        # 2. 筛选实时交易
        realtime_txs, backfill_txs = self.filter_realtime(all_txs)
        
        # 3. 计算统计
        stats = self.calculate_stats(realtime_txs)
        
        # 4. 生成报告
        return {
            'summary': {
                'total_transactions': len(all_txs),
                'realtime_transactions': len(realtime_txs),
                'backfill_transactions': len(backfill_txs),
                'realtime_ratio': f"{len(realtime_txs) / len(all_txs) * 100:.1f}%"
            },
            'delay_stats': stats,
            'realtime_txs': realtime_txs,
            'backfill_txs': backfill_txs
        }
    
    def print_report(self, report):
        """打印报告"""
        
        if 'error' in report:
            print(f"\n❌ 错误: {report['error']}\n")
            return
        
        print("\n" + "="*70)
        print("跟踪效果验证报告")
        print("="*70)
        
        # 摘要
        summary = report['summary']
        print(f"\n📊 摘要:")
        print(f"  总交易数: {summary['total_transactions']}笔")
        print(f"  实时捕获: {summary['realtime_transactions']}笔 ({summary['realtime_ratio']})")
        print(f"  离线回溯: {summary['backfill_transactions']}笔")
        
        # 延迟分析
        stats = report['delay_stats']
        if stats:
            print(f"\n⏱️ 延迟分析（实时交易）:")
            print(f"  平均延迟: {stats['avg_delay']:.2f}秒")
            print(f"  中位延迟: {stats['median_delay']:.2f}秒")
            print(f"  最小延迟: {stats['min_delay']:.2f}秒")
            print(f"  最大延迟: {stats['max_delay']:.2f}秒")
            print(f"  P95延迟: {stats['p95_delay']:.2f}秒")
            
            # 延迟分布
            print(f"\n📈 延迟分布:")
            dist = stats['distribution']
            for range_name, count in dist.items():
                percentage = count / stats['count'] * 100
                bar = '█' * int(percentage / 2)
                print(f"  {range_name:8} | {bar} {count}笔 ({percentage:.1f}%)")
            
            # 结论
            print(f"\n✅ 结论:")
            avg = stats['avg_delay']
            if avg < 5:
                print(f"  🟢 优秀！平均延迟 {avg:.2f}秒")
            elif avg < 8:
                print(f"  🟡 良好！平均延迟 {avg:.2f}秒")
            elif avg < 15:
                print(f"  🟠 一般！平均延迟 {avg:.2f}秒，建议优化")
            else:
                print(f"  🔴 较差！平均延迟 {avg:.2f}秒，需要优化")
        
        print("="*70 + "\n")


# 使用示例
if __name__ == '__main__':
    import sys
    
    # 配置
    EXCEL_FILE = r"C:\Users\pill\Desktop\跟踪聪明钱\5.0ing 查询改异步，再次重构\database\transactions_CyaE1V_a54o.xlsx"
    API_KEY = os.getenv('HELIUS_API_KEY')
    
    if not API_KEY:
        print("❌ 错误: 未找到 HELIUS_API_KEY")
        print("请确保 .env 文件存在且包含 HELIUS_API_KEY=your_key")
        print(f"\n当前目录: {os.getcwd()}")
        print(f".env 文件是否存在: {os.path.exists('.env')}")
        sys.exit(1)
    
    print(f"✅ 成功读取 API Key: {API_KEY[:10]}...")
    
    # 创建验证器
    validator = SimpleTrackingValidator(EXCEL_FILE, API_KEY)
    
    # 执行验证
    print("🚀 开始验证跟踪效果...")
    print("⚠️  这将查询所有交易的链上时间（可能需要几分钟）\n")
    
    report = validator.validate()
    
    # 打印报告
    validator.print_report(report)
